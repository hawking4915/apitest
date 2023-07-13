# -*- coding: utf-8 -*-
# @Time : 2019/11/13 9:31
# @Author : kira
# @Email : 262667641@qq.com
# @File : do_excel.py
# @Project : risk_project
import sys

sys.path.append("../../")
sys.path.append("../../common")
from openpyxl import load_workbook
from common.utils.singleton import singleton
from common.file_handling import logger


@singleton
class DoExcel:
    
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)
        self.init_sheet = self.wb["init"]
    
    # def __enter__(self):
    #     self.wb = load_workbook(self.file_name)
    #     self.init_sheet = self.wb['init']
    #     return self
    #
    # def __exit__(self, exc_type, exc_val, exc_tb):
    #     self.wb.save(self.file_name)
    #     self.wb.close()
    
    def do_excel_yield(self):
        """
        读取excel数据
        Returns:
    
        """
        sheets = eval(self.get_excel_init().get("sheets"))
        for sheet_name in sheets:
            sheet = self.wb[sheet_name]
            max_row = sheet.max_row
            max_column = sheet.max_column
            first_header = []
            for i in range(1, max_column + 1):
                first_header.append(sheet.cell(1, i).value)
            for i in range(2, max_row + 1):
                sub_data = {}
                for k in range(1, max_column + 1):
                    sub_data[first_header[k - 1]] = sheet.cell(i, k).value
                    sub_data["sheet"] = sheet_name
                yield sub_data
    
    @logger.log_decorator()
    def write_back(self, sheet_name, i, **kwargs):
        """
    
        Args:
            sheet_name:sheet 名称
            i:序号
            response_value: 响应结果
            test_result: 测试结果
            assert_log: 报错结果
        Returns:
        """
        response_value = kwargs.get("response")
        test_result = kwargs.get("test_result")
        assert_log = kwargs.get("assert_log")
        sheet = self.wb[sheet_name]
        sheet.cell(i + 1, 24).value = response_value
        sheet.cell(i + 1, 25).value = test_result
        sheet.cell(i + 1, 26).value = assert_log
        self.wb.save(self.file_name)
    
    @logger.log_decorator()
    def clear_date(self):
        """
        执行清空单元格数据
        Returns:
    
        """
        sheets = eval(self.get_excel_init().get("sheets"))
        
        for sheet_name in sheets:
            sheet = self.wb[sheet_name]
            max_row = sheet.max_row  # 获取最大行
            for i in range(2, max_row + 1):
                sheet.cell(i, 24).value = ""
                sheet.cell(i, 25).value = ""
                sheet.cell(i, 26).value = ""
        self.wb.save(self.file_name)
        return f"清空指定 {sheets} 中的单元格成功"
    
    @logger.log_decorator()
    def get_excel_init(self):
        """
        获取 excel 中 sheet 名称为 init 中的基础数据
        Returns:init 表中的所有数据
        """
        max_row = self.init_sheet.max_row  # 获取最大行
        max_column = self.init_sheet.max_column  # 获取最大列
        first_head = []  # 存储标题的 list
        for i in range(max_column):
            first_head.append(self.init_sheet.cell(1, i + 1).value)
        init = {}
        for k in range(2, max_row + 1):
            for i, v in enumerate(first_head):
                init[v] = self.init_sheet.cell(k, i + 1).value
            if init.get("run").upper() == "YES":
                break
        return init
    
    def get_excel_init_and_cases(self):
        """
    
        Returns:初始化数据及测试用例组成的元组
    
        """
        try:
            self.clear_date()
            test_case = self.do_excel_yield()
            init_data = self.get_excel_init()
        except Exception as e:
            raise e
        return init_data, test_case
    
    def close_excel(self):
        self.wb.close()


if __name__ == '__main__':
    from common.config import Config
    
    file_n = Config.test_case
    excel = DoExcel(file_n)
# excel.get_excel_init()
# excel.do_excel()
# excel.clear_date()
# excel.do_excel()
